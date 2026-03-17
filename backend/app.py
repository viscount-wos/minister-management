from flask import Flask, request, jsonify, send_from_directory, Response
from flask_cors import CORS
import os
import hashlib
import time as time_module
import requests as http_requests
from dotenv import load_dotenv
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

# Load environment variables FIRST
load_dotenv()

from database import (
    init_db, get_all_players, get_player_by_fid,
    save_player, delete_player, calculate_points, get_db,
    get_research_day, get_show_fire_crystals, set_setting, get_setting,
    get_time_preference_counts
)
import json

STATIC_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static')

app = Flask(__name__, static_folder=None)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'dev-secret-key')
CORS(app)

# Admin credentials
ADMIN_PASSWORD = os.getenv('ADMIN_PASSWORD', 'admin123')
MINISTER_PASSWORD = os.getenv('MINISTER_PASSWORD', 'minister123')

# Initialize database (registers teardown handler)
init_db(app)

# Health check endpoint
@app.route('/health', methods=['GET'])
def health_check():
    return jsonify({'status': 'healthy'}), 200

# Serve React app - handles SPA routing for all non-API paths
@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def serve(path):
    if path and os.path.exists(os.path.join(STATIC_DIR, path)):
        return send_from_directory(STATIC_DIR, path)
    return send_from_directory(STATIC_DIR, 'index.html')

# API Routes

@app.route('/api/player/submit', methods=['POST'])
def submit_player():
    """Submit or update player information."""
    try:
        data = request.json

        # Validate required fields
        required_fields = ['fid', 'game_name', 'alliance']
        for field in required_fields:
            if field not in data or not str(data[field]).strip():
                return jsonify({'error': f'Missing required field: {field}'}), 400

        # Extract time slots (supports both legacy list and per-day dict)
        time_slots_by_day = data.pop('time_slots_by_day', None)
        time_slots = data.pop('time_slots', [])
        if time_slots_by_day:
            time_slots = time_slots_by_day  # pass dict to save_player

        # Set defaults for numeric fields
        numeric_fields = [
            'construction_speedups_days', 'research_speedups_days',
            'troop_training_speedups_days', 'general_speedups_days',
            'fire_crystals', 'refined_fire_crystals', 'fire_crystal_shards'
        ]
        for field in numeric_fields:
            if field not in data:
                data[field] = 0
            else:
                data[field] = float(data[field]) if '.' in str(data[field]) or 'speedups' in field else int(data[field])

        # Save player
        player_id = save_player(data, time_slots)

        return jsonify({
            'success': True,
            'player_id': player_id,
            'message': 'Player information saved successfully'
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/player/check-duplicate', methods=['POST'])
def check_duplicate():
    """Check if a player with the given FID or game name already exists."""
    try:
        data = request.json
        fid = data.get('fid', '').strip()
        game_name = data.get('game_name', '').strip()

        db = get_db()
        cursor = db.cursor()
        result = {'fid_exists': False, 'name_exists': False}

        if fid:
            cursor.execute('SELECT id FROM players WHERE fid = ?', (fid,))
            if cursor.fetchone():
                result['fid_exists'] = True

        if game_name:
            cursor.execute('SELECT id FROM players WHERE LOWER(game_name) = LOWER(?)', (game_name,))
            if cursor.fetchone():
                result['name_exists'] = True

        return jsonify(result), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/player/<fid>', methods=['GET'])
def get_player(fid):
    """Get player information by FID."""
    try:
        player = get_player_by_fid(fid)
        if player:
            return jsonify(player), 200
        return jsonify({'error': 'Player not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/player/wos-lookup', methods=['POST'])
def wos_lookup():
    """Look up player info from the WOS game API."""
    try:
        data = request.json
        fid = data.get('fid', '').strip()

        if not fid:
            return jsonify({'error': 'FID is required'}), 400

        # Build signed request for WOS API
        secret = 'tB87#kPtkxqOS2'
        ts = str(int(time_module.time() * 1e9))
        form_data = f'fid={fid}&time={ts}'
        sign = hashlib.md5((form_data + secret).encode()).hexdigest()
        body = f'sign={sign}&{form_data}'

        response = http_requests.post(
            'https://wos-giftcode-api.centurygame.com/api/player',
            data=body,
            headers={
                'Content-Type': 'application/x-www-form-urlencoded',
                'Accept': 'application/json',
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'Origin': 'https://wos-giftcode.centurygame.com',
                'Referer': 'https://wos-giftcode.centurygame.com/',
            },
            timeout=10
        )

        result = response.json()

        if result.get('code') != 0:
            return jsonify({'error': 'Player not found in WOS'}), 404

        wos_data = result['data']
        return jsonify({
            'success': True,
            'fid': str(wos_data['fid']),
            'nickname': wos_data.get('nickname', ''),
            'kid': wos_data.get('kid'),
            'stove_lv': wos_data.get('stove_lv'),
            'stove_lv_content': wos_data.get('stove_lv_content', ''),
            'avatar_image': wos_data.get('avatar_image', ''),
        }), 200

    except http_requests.exceptions.Timeout:
        return jsonify({'error': 'WOS API timed out'}), 504
    except http_requests.exceptions.RequestException as e:
        return jsonify({'error': f'Failed to reach WOS API: {str(e)}'}), 502
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/settings/research-day', methods=['GET'])
def get_research_day_setting():
    """Get the current research day setting (public endpoint)."""
    return jsonify({'research_day': get_research_day()}), 200

@app.route('/api/admin/settings/research-day', methods=['PUT'])
def set_research_day_setting():
    """Set the research day to 'tuesday' or 'friday'."""
    try:
        auth_header = request.headers.get('Authorization')
        if not auth_header or 'token' not in auth_header:
            return jsonify({'error': 'Unauthorized'}), 401

        data = request.json
        day = data.get('research_day', '').lower()
        if day not in ('tuesday', 'friday'):
            return jsonify({'error': 'Invalid value. Must be "tuesday" or "friday"'}), 400

        set_setting('research_day', day)
        return jsonify({'success': True, 'research_day': day}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/settings/show-fire-crystals', methods=['GET'])
def get_fire_crystals_setting():
    """Get whether fire crystal fields should be shown (public endpoint)."""
    return jsonify({'show_fire_crystals': get_show_fire_crystals()}), 200

@app.route('/api/admin/settings/show-fire-crystals', methods=['PUT'])
def set_fire_crystals_setting():
    """Toggle fire crystal fields visibility."""
    try:
        auth_header = request.headers.get('Authorization')
        if not auth_header or 'token' not in auth_header:
            return jsonify({'error': 'Unauthorized'}), 401

        data = request.json
        show = data.get('show_fire_crystals', False)
        set_setting('show_fire_crystals', 'true' if show else 'false')
        return jsonify({'success': True, 'show_fire_crystals': show}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/login', methods=['POST'])
def admin_login():
    """Authenticate admin or minister user."""
    try:
        data = request.json
        password = data.get('password', '')

        if password == ADMIN_PASSWORD:
            return jsonify({
                'success': True,
                'role': 'admin',
                'token': 'admin-token'
            }), 200
        elif password == MINISTER_PASSWORD:
            return jsonify({
                'success': True,
                'role': 'minister',
                'token': 'minister-token'
            }), 200
        else:
            return jsonify({'error': 'Invalid password'}), 401

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/players', methods=['GET'])
def get_players():
    """Get all players with calculated points."""
    try:
        # Simple auth check
        auth_header = request.headers.get('Authorization')
        if not auth_header or 'token' not in auth_header:
            return jsonify({'error': 'Unauthorized'}), 401

        players = get_all_players()

        # Add calculated points for each day
        research_day = get_research_day()
        for player in players:
            player['monday_points'] = calculate_points(player, 'monday')
            player['research_points'] = calculate_points(player, research_day)
            player['thursday_points'] = calculate_points(player, 'thursday')
            player['research_day'] = research_day

        return jsonify(players), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/player/<int:player_id>', methods=['PUT'])
def update_player(player_id):
    """Update player information."""
    try:
        # Simple auth check
        auth_header = request.headers.get('Authorization')
        if not auth_header or 'token' not in auth_header:
            return jsonify({'error': 'Unauthorized'}), 401

        data = request.json
        time_slots_by_day = data.pop('time_slots_by_day', None)
        time_slots = data.pop('time_slots', [])
        if time_slots_by_day:
            time_slots = time_slots_by_day

        # Update player
        save_player(data, time_slots)

        return jsonify({'success': True, 'message': 'Player updated'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/player/<int:player_id>', methods=['DELETE'])
def remove_player(player_id):
    """Delete a player."""
    try:
        # Simple auth check
        auth_header = request.headers.get('Authorization')
        if not auth_header or 'token' not in auth_header:
            return jsonify({'error': 'Unauthorized'}), 401

        delete_player(player_id)
        return jsonify({'success': True, 'message': 'Player deleted'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/assignments/auto-assign', methods=['POST'])
def auto_assign():
    """Auto-assign players to time slots based on points."""
    try:
        # Simple auth check
        auth_header = request.headers.get('Authorization')
        if not auth_header or 'token' not in auth_header:
            return jsonify({'error': 'Unauthorized'}), 401

        data = request.json
        day = data.get('day', '').lower()

        research_day = get_research_day()
        valid_days = ['monday', research_day, 'thursday']
        if day not in valid_days:
            return jsonify({'error': 'Invalid day'}), 400

        players = get_all_players()

        # Map day to day_type for time preferences
        day_type_map = {'monday': 'construction', 'thursday': 'troop'}
        # Research day (tuesday or friday) maps to 'research'
        day_type_map[research_day] = 'research'
        day_type = day_type_map.get(day, 'construction')

        # Calculate points for this day
        for player in players:
            player['points'] = calculate_points(player, day)

        # Sort by points (descending)
        players.sort(key=lambda p: p['points'], reverse=True)

        # Generate 30-minute time slots starting at 23:50 (previous day)
        # through 23:50+ (end of day), covering ~24.5 hours.
        # Slots: 23:50, 00:20, 00:50, 01:20, ..., 23:20, 23:50+
        time_slots = ['23:50']
        hour, minute = 0, 20
        while True:
            slot = f"{hour:02d}:{minute:02d}"
            if slot == '23:50':
                time_slots.append('23:50+')
                break
            time_slots.append(slot)
            minute += 30
            if minute >= 60:
                minute -= 60
                hour += 1

        # Fetch sticky assignments BEFORE clearing
        db = get_db()
        cursor = db.cursor()
        cursor.execute('''
            SELECT a.player_id, a.time_slot, p.fid, p.game_name,
                   p.avatar_image, p.stove_lv, p.stove_lv_content, p.alliance,
                   p.construction_speedups_days, p.research_speedups_days,
                   p.troop_training_speedups_days, p.general_speedups_days,
                   p.fire_crystals, p.refined_fire_crystals, p.fire_crystal_shards
            FROM assignments a
            JOIN players p ON a.player_id = p.id
            WHERE a.day = ? AND a.is_sticky = 1
        ''', (day,))
        sticky_rows = cursor.fetchall()
        sticky_slots = {}  # time_slot -> player data
        sticky_player_ids = set()
        for row in sticky_rows:
            row_dict = dict(row)
            row_dict['points'] = calculate_points(row_dict, day)
            sticky_slots[row['time_slot']] = {
                'id': row['player_id'],
                'player_id': row['player_id'],
                'fid': row['fid'],
                'game_name': row['game_name'],
                'points': row_dict['points'],
                'avatar_image': row_dict.get('avatar_image') or '',
                'stove_lv': row_dict.get('stove_lv'),
                'stove_lv_content': row_dict.get('stove_lv_content') or '',
                'alliance': row_dict.get('alliance') or '',
                'is_sticky': True,
            }
            sticky_player_ids.add(row['player_id'])

        # Assignment logic
        assignments = {slot: [] for slot in time_slots}
        unassigned = []

        # Pre-fill sticky assignments
        for slot, player_data in sticky_slots.items():
            if slot in assignments:
                assignments[slot].append(player_data)

        for player in players:
            # Skip players that are sticky-assigned
            if player['id'] in sticky_player_ids:
                continue

            # Use day-specific time preferences
            time_slots_by_day = player.get('time_slots_by_day', {})
            player_time_prefs = set(time_slots_by_day.get(day_type, player.get('time_slots', [])))

            # Find matching 30-min slots for player's hourly preferences
            # With ±20 min tolerance, each hour H maps to 3 slots:
            #   (H-1):50  — starts 10 min before the hour (within 20 min)
            #   H:20      — starts 20 min after the hour (within 20 min)
            #   H:50      — within the selected hour
            matching_slots = []
            for pref in player_time_prefs:
                if ':' in pref:
                    h = int(pref.split(':')[0])
                    prev_h = (h - 1) % 24
                    # Previous hour's :50 slot
                    if h == 0:
                        matching_slots.append('23:50')  # The pre-midnight slot
                    else:
                        matching_slots.append(f"{prev_h:02d}:50")
                    # Current hour's :20 and :50 slots
                    matching_slots.append(f"{h:02d}:20")
                    # For hour 23, the :50 slot is "23:50+" (end of day),
                    # NOT "23:50" which is the pre-midnight slot (previous day)
                    if h == 23:
                        matching_slots.append('23:50+')
                    else:
                        matching_slots.append(f"{h:02d}:50")

            assigned = False
            for slot in matching_slots:
                if slot in assignments and len(assignments[slot]) == 0:
                    assignments[slot].append({
                        'id': player['id'],
                        'player_id': player['id'],
                        'fid': player['fid'],
                        'game_name': player['game_name'],
                        'points': player['points'],
                        'avatar_image': player.get('avatar_image', ''),
                        'stove_lv': player.get('stove_lv'),
                        'stove_lv_content': player.get('stove_lv_content', ''),
                        'alliance': player.get('alliance', ''),
                        'is_sticky': False,
                    })
                    assigned = True
                    break

            if not assigned:
                unassigned.append({
                    'id': player['id'],
                    'player_id': player['id'],
                    'fid': player['fid'],
                    'game_name': player['game_name'],
                    'points': player['points'],
                    'preferred_times': list(player_time_prefs),
                    'avatar_image': player.get('avatar_image', ''),
                    'stove_lv': player.get('stove_lv'),
                    'stove_lv_content': player.get('stove_lv_content', ''),
                    'alliance': player.get('alliance', ''),
                    'is_sticky': False,
                })

        # Clear existing non-sticky assignments for this day, then clear sticky too (we'll re-insert all)
        cursor.execute('DELETE FROM assignments WHERE day = ?', (day,))

        # Save new assignments (including sticky ones)
        for time_slot, slot_players in assignments.items():
            for position, player in enumerate(slot_players):
                cursor.execute('''
                    INSERT INTO assignments (player_id, day, time_slot, position, is_assigned, is_sticky)
                    VALUES (?, ?, ?, ?, 1, ?)
                ''', (player['id'], day, time_slot, position, 1 if player.get('is_sticky') else 0))

        db.commit()

        return jsonify({
            'success': True,
            'assignments': assignments,
            'unassigned': unassigned
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/assignments/<day>', methods=['GET'])
def get_assignments(day):
    """Get assignments for a specific day."""
    try:
        # Simple auth check
        auth_header = request.headers.get('Authorization')
        if not auth_header or 'token' not in auth_header:
            return jsonify({'error': 'Unauthorized'}), 401

        db = get_db()
        cursor = db.cursor()
        cursor.execute('''
            SELECT
                a.id, a.time_slot, a.position, a.is_assigned, a.is_sticky,
                p.id as player_id, p.fid, p.game_name,
                p.construction_speedups_days, p.research_speedups_days,
                p.troop_training_speedups_days, p.general_speedups_days,
                p.fire_crystals, p.refined_fire_crystals, p.fire_crystal_shards,
                p.avatar_image, p.stove_lv, p.stove_lv_content, p.alliance
            FROM assignments a
            JOIN players p ON a.player_id = p.id
            WHERE a.day = ?
            ORDER BY a.time_slot, a.position
        ''', (day.lower(),))

        rows = cursor.fetchall()
        assignments = {}
        for row in rows:
            slot = row['time_slot']
            if slot not in assignments:
                assignments[slot] = []
            row_dict = dict(row)
            row_dict['points'] = calculate_points(row_dict, day.lower())
            assignments[slot].append(row_dict)

        return jsonify(assignments), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/assignments/update', methods=['POST'])
def update_assignments():
    """Update assignments after drag-and-drop."""
    try:
        # Simple auth check
        auth_header = request.headers.get('Authorization')
        if not auth_header or 'token' not in auth_header:
            return jsonify({'error': 'Unauthorized'}), 401

        data = request.json
        day = data.get('day')
        assignments = data.get('assignments', {})

        db = get_db()
        cursor = db.cursor()

        # Clear existing assignments
        cursor.execute('DELETE FROM assignments WHERE day = ?', (day,))

        # Save new assignments (enforce max 1 player per slot)
        for time_slot, slot_players in assignments.items():
            if slot_players:
                player = slot_players[0]  # Only take first player per slot
                cursor.execute('''
                    INSERT INTO assignments (player_id, day, time_slot, position, is_assigned, is_sticky)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (player['player_id'], day, time_slot, 0,
                      player.get('is_assigned', True),
                      1 if player.get('is_sticky') else 0))

        db.commit()

        return jsonify({'success': True}), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/export', methods=['GET'])
def export_assignments():
    """Export assignments for all days to a single Excel workbook."""
    try:
        # Simple auth check
        auth_header = request.headers.get('Authorization')
        if not auth_header or 'token' not in auth_header:
            return jsonify({'error': 'Unauthorized'}), 401

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        separator_fill = PatternFill(start_color="F4B942", end_color="F4B942", fill_type="solid")
        separator_font = Font(bold=True, color="000000")

        headers = [
            'Time Slot', 'FID', 'Alliance', 'Game Name',
            'Construction (days)', 'Research (days)',
            'Troop Training (days)', 'General (days)',
            'Fire Crystals', 'Refined Fire Crystals',
            'Crystal Shards', 'Points'
        ]

        col_widths = [15, 15, 10, 25, 18, 15, 20, 15, 13, 18, 14, 12]

        research_day = get_research_day()
        research_label = 'Tuesday - Research' if research_day == 'tuesday' else 'Friday - Research'
        days = [
            ('monday', 'Monday - Construction'),
            (research_day, research_label),
            ('thursday', 'Thursday - Troop Training'),
        ]

        db = get_db()
        cursor = db.cursor()

        for day_key, day_title in days:
            ws = wb.create_sheet(title=day_title)
            ws.append(headers)

            # Style headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # Get assigned players
            cursor.execute('''
                SELECT a.time_slot, p.*
                FROM assignments a
                JOIN players p ON a.player_id = p.id
                WHERE a.day = ? AND a.is_assigned = 1
                ORDER BY a.time_slot, a.position
            ''', (day_key,))

            assigned_rows = cursor.fetchall()
            assigned_player_ids = set()

            for row in assigned_rows:
                player = dict(row)
                assigned_player_ids.add(player['id'])
                points = calculate_points(player, day_key)
                ws.append([
                    '23:50 (+1d)' if row['time_slot'] == '23:50+' else row['time_slot'],
                    player['fid'],
                    player.get('alliance', ''),
                    player['game_name'],
                    player['construction_speedups_days'],
                    player['research_speedups_days'],
                    player['troop_training_speedups_days'],
                    player['general_speedups_days'],
                    player['fire_crystals'],
                    player['refined_fire_crystals'],
                    player['fire_crystal_shards'],
                    points,
                ])

            # Get all players to find unassigned ones
            cursor.execute('SELECT * FROM players ORDER BY id')
            all_players = [dict(r) for r in cursor.fetchall()]
            unassigned = [p for p in all_players if p['id'] not in assigned_player_ids]

            # Sort unassigned by points descending
            for p in unassigned:
                p['_points'] = calculate_points(p, day_key)
            unassigned.sort(key=lambda p: p['_points'], reverse=True)

            if unassigned:
                # Separator row
                sep_row_num = ws.max_row + 2  # skip a blank row
                ws.append([])  # blank row
                ws.append(['UNASSIGNED PLAYERS'] + [''] * (len(headers) - 1))
                for cell in ws[sep_row_num]:
                    cell.fill = separator_fill
                    cell.font = separator_font

                for player in unassigned:
                    ws.append([
                        'Unassigned',
                        player['fid'],
                        player.get('alliance', ''),
                        player['game_name'],
                        player['construction_speedups_days'],
                        player['research_speedups_days'],
                        player['troop_training_speedups_days'],
                        player['general_speedups_days'],
                        player['fire_crystals'],
                        player['refined_fire_crystals'],
                        player['fire_crystal_shards'],
                        player['_points'],
                    ])

            # Set column widths
            for i, width in enumerate(col_widths):
                ws.column_dimensions[chr(65 + i)].width = width

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'ministry_assignments_{datetime.now().strftime("%Y%m%d")}.xlsx'
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/settings/publish', methods=['PUT'])
def publish_schedule():
    """Publish assignments for a day so they appear on the public page.
    Supports multiple days — stores as comma-separated list.
    """
    try:
        auth_header = request.headers.get('Authorization')
        if not auth_header or 'token' not in auth_header:
            return jsonify({'error': 'Unauthorized'}), 401

        data = request.json
        day = data.get('day', '').lower()
        research_day = get_research_day()
        valid_days = ['monday', research_day, 'thursday']
        if day not in valid_days:
            return jsonify({'error': 'Invalid day'}), 400

        # Add to existing published days
        current = get_setting('published_days', '')
        days_set = set(d for d in current.split(',') if d)
        days_set.add(day)
        set_setting('published_days', ','.join(sorted(days_set)))
        return jsonify({'success': True, 'published_days': sorted(days_set)}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/settings/unpublish', methods=['PUT'])
def unpublish_schedule():
    """Remove a specific day from published schedules."""
    try:
        auth_header = request.headers.get('Authorization')
        if not auth_header or 'token' not in auth_header:
            return jsonify({'error': 'Unauthorized'}), 401

        data = request.json
        day = data.get('day', '').lower()

        current = get_setting('published_days', '')
        days_set = set(d for d in current.split(',') if d)
        days_set.discard(day)
        set_setting('published_days', ','.join(sorted(days_set)))
        return jsonify({'success': True, 'published_days': sorted(days_set)}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/published-schedule/<day>', methods=['GET'])
def get_published_schedule(day):
    """Get the published schedule for a specific day (public endpoint).
    Returns only player name, alliance, and time slot — no points or resources.
    """
    try:
        published_days = get_setting('published_days', '')
        days_list = [d for d in published_days.split(',') if d]
        if day.lower() not in days_list:
            return jsonify({'published': False}), 200

        db = get_db()
        cursor = db.cursor()
        cursor.execute('''
            SELECT a.time_slot, p.game_name, p.alliance
            FROM assignments a
            JOIN players p ON a.player_id = p.id
            WHERE a.day = ? AND a.is_assigned = 1
            ORDER BY a.time_slot, a.position
        ''', (day.lower(),))

        rows = cursor.fetchall()
        assignments = {}
        for row in rows:
            slot = row['time_slot']
            if slot not in assignments:
                assignments[slot] = []
            assignments[slot].append({
                'game_name': row['game_name'],
                'alliance': row['alliance'] or '',
            })

        day_labels = {
            'monday': 'Monday - Construction',
            'tuesday': 'Tuesday - Research',
            'friday': 'Friday - Research',
            'thursday': 'Thursday - Troop Training',
        }

        return jsonify({
            'published': True,
            'day': day.lower(),
            'day_label': day_labels.get(day.lower(), day),
            'assignments': assignments,
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/settings/published-days', methods=['GET'])
def get_published_days_setting():
    """Get which days are currently published (public endpoint)."""
    published_days = get_setting('published_days', '')
    days_list = [d for d in published_days.split(',') if d]
    return jsonify({'published_days': days_list}), 200


@app.route('/api/time-preferences/heatmap', methods=['GET'])
def get_heatmap():
    """Get time preference counts per slot per day_type (public endpoint)."""
    try:
        counts = get_time_preference_counts()
        return jsonify(counts), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/player/<fid>/assignments', methods=['GET'])
def get_player_assignments(fid):
    """Get a player's current assignments by FID (public endpoint).
    Returns only day and time_slot — no points or resource data.
    """
    try:
        player = get_player_by_fid(fid)
        if not player:
            return jsonify({'error': 'Player not found'}), 404

        db = get_db()
        cursor = db.cursor()
        cursor.execute('''
            SELECT day, time_slot
            FROM assignments
            WHERE player_id = ? AND is_assigned = 1
            ORDER BY day, time_slot
        ''', (player['id'],))

        assignments = {}
        for row in cursor.fetchall():
            day = row['day']
            if day not in assignments:
                assignments[day] = []
            assignments[day].append({'time_slot': row['time_slot']})

        return jsonify({'assignments': assignments}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/players/export-json', methods=['GET'])
def export_players_json():
    """Export all players as JSON (admin auth required)."""
    try:
        auth_header = request.headers.get('Authorization')
        if not auth_header or 'token' not in auth_header:
            return jsonify({'error': 'Unauthorized'}), 401

        players = get_all_players()

        # Clean up internal fields for export
        export_players = []
        for p in players:
            export_players.append({
                'fid': p['fid'],
                'game_name': p['game_name'],
                'alliance': p.get('alliance', ''),
                'construction_speedups_days': p['construction_speedups_days'],
                'research_speedups_days': p['research_speedups_days'],
                'troop_training_speedups_days': p['troop_training_speedups_days'],
                'general_speedups_days': p['general_speedups_days'],
                'fire_crystals': p['fire_crystals'],
                'refined_fire_crystals': p['refined_fire_crystals'],
                'fire_crystal_shards': p['fire_crystal_shards'],
                'avatar_image': p.get('avatar_image', ''),
                'stove_lv': p.get('stove_lv'),
                'stove_lv_content': p.get('stove_lv_content', ''),
                'timezone': p.get('timezone', ''),
                'time_slots_by_day': p.get('time_slots_by_day', {}),
            })

        export_data = {
            'version': 1,
            'exported_at': datetime.now().isoformat(),
            'players': export_players,
        }

        output = json.dumps(export_data, indent=2)
        filename = f'players_backup_{datetime.now().strftime("%Y%m%d")}.json'
        return Response(
            output,
            mimetype='application/json',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/players/import', methods=['POST'])
def import_players_json():
    """Import players from JSON (admin auth required). Upserts by FID."""
    try:
        auth_header = request.headers.get('Authorization')
        if not auth_header or 'token' not in auth_header:
            return jsonify({'error': 'Unauthorized'}), 401

        data = request.json
        if not data or 'players' not in data:
            return jsonify({'error': 'Invalid format: expected {players: [...]}'}), 400

        imported = 0
        updated = 0
        errors = 0

        for p in data['players']:
            try:
                fid = p.get('fid', '').strip()
                if not fid:
                    errors += 1
                    continue

                # Check if player exists
                existing = get_player_by_fid(fid)

                player_data = {
                    'fid': fid,
                    'game_name': p.get('game_name', 'Unknown'),
                    'alliance': p.get('alliance', ''),
                    'construction_speedups_days': float(p.get('construction_speedups_days', 0)),
                    'research_speedups_days': float(p.get('research_speedups_days', 0)),
                    'troop_training_speedups_days': float(p.get('troop_training_speedups_days', 0)),
                    'general_speedups_days': float(p.get('general_speedups_days', 0)),
                    'fire_crystals': int(p.get('fire_crystals', 0)),
                    'refined_fire_crystals': int(p.get('refined_fire_crystals', 0)),
                    'fire_crystal_shards': int(p.get('fire_crystal_shards', 0)),
                    'avatar_image': p.get('avatar_image', ''),
                    'stove_lv': p.get('stove_lv'),
                    'stove_lv_content': p.get('stove_lv_content', ''),
                    'timezone': p.get('timezone', ''),
                }

                time_slots = p.get('time_slots_by_day', p.get('time_slots', []))
                save_player(player_data, time_slots)

                if existing:
                    updated += 1
                else:
                    imported += 1
            except Exception:
                errors += 1

        return jsonify({
            'success': True,
            'imported': imported,
            'updated': updated,
            'errors': errors,
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/players/delete-all', methods=['DELETE'])
def delete_all_players():
    """Delete all players and their assignments."""
    try:
        auth_header = request.headers.get('Authorization')
        if not auth_header or 'token' not in auth_header:
            return jsonify({'error': 'Unauthorized'}), 401

        db = get_db()
        cursor = db.cursor()
        cursor.execute('DELETE FROM assignments')
        cursor.execute('DELETE FROM time_preferences')
        cursor.execute('DELETE FROM players')
        db.commit()

        return jsonify({'success': True, 'message': 'All players deleted'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    port = int(os.getenv('PORT', 8080))
    app.run(host='0.0.0.0', port=port, debug=os.getenv('FLASK_ENV') == 'development')
